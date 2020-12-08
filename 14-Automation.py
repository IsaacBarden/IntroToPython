#Spreadsheet processing
import openpyxl as xl
from openpyxl.chart import BarChart, Reference

def process_spreadsheet(filename):
    wb = xl.load_workbook("IntroToPython\\Files\\" + filename)
    sheet = wb["Sheet1"]

    for row in range(2, sheet.max_row + 1):
        cell = sheet.cell(row, 3)
        reduced_price = cell.value * 0.9
        reduced_price_cell = sheet.cell(row, 4)
        reduced_price_cell.value = reduced_price

    values = Reference(sheet, min_row=2, max_row=sheet.max_row, min_col=4, max_col=4)
    chart = BarChart()
    chart.add_data(values)
    sheet.add_chart(chart, "E2")


    wb.save("IntroToPython\\Files\\OutputExSpreadsheet.xlsx")

    
filename = "ExSpreadsheet.xlsx"
process_spreadsheet(filename)